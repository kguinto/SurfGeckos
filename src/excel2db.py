"""
Currently creates a databse of many but not all tables. Currently fails on F-1, H,J

H and J just need the schema filled in.

F-1 is very different from the rest and may need a new function.

From command line:
python excel2db.py <excel file> 
OR
python excel2db.py <excel file> --mongo

The former will created a sqlite database surfer.db.
The latter will work if mongodb is running and create a mongo db named surfer.

From within your own code

from excel2db import Loader
myfile = <excel file>
mydb = <database name>
Loader(myfile, mydb)
OR
Loader(myfile, mydb, mongo=True)



"""

########
# Table F-1 is very different and will need to be handled specially

###
import sys
import json
import argparse
import logging

import pandas

import openpyxl

logging_format = '%(asctime)s %(message)s'
logging.basicConfig(format=logging_format, level=logging.INFO)

# TODO: Explore sqlalchemy
# TODO: Make schema for all files

logging.debug("Begin")

##############################
# Establish table parameters #
##############################

A = ['contaminant', 'soil_far', 'groundwater_far', 'soil_close', 'groundwater_close']
C = ['contaminant', 'state_A', 'state_B', 'indoor_residential', 'indoor_commerical', 'shallow_residential', 'shallow_commercial']
D = ['contaminant', 'freshwater', 'marine', 'estuarine']
A1 = ['contaminant', 'eal', 'basis', 'gross_contamination', 'toxicity', 'background', 'direct_exposure', 'vapor_intrusion', 'drinking_water']
C1a = ['contaminant', 'extra', 'state_A', 'state_B', 'unrestricted', 'commercial']
C2 = ['contaminant', 'state_A', 'state_B', 'unrestricted_lowest', 'unrestricted_carcinogenic', 'unrestricted_noncarcinogenic', 'commercial_lowest', 'commercial_carcinogenic', 'commercial_noncarcinogenic']
C3 =  ['contaminant', 'state_A', 'state_B', 'urf', 'rfc', 'unrestricted_lowest', 'unrestricted_carcinogenic', 'unrestricted_noncarcinogenic', 'commercial_lowest', 'commercial_carcinogenic', 'commercial_noncarcinogenic', 'odor_threshold']
D1a = ['contaminant', 'eal', 'basis', 'gross_contamination', 'toxicity', 'vapor_intrusion', 'aquatic_impacts']
D1c = ['contaminant', 'eal', 'basis', 'gross_contamination', 'vapor_intrusion', 'aquatic_impacts']
D2a = ['containant', 'eal', 'basis', 'gross_contamination', 'toxicity', 'aquatic_impacts', 'biaccumulation']
D2b = ['contaminant', 'eal', 'basis', 'gross_contamination', 'toxicity', 'biaccumulation']
D3a = ['contaminant', 'eal', 'basis', 'hdoh', 'other_criteria', 'reference', 'risk_action_level', 'risk_basis']
D3b = ['contaminant', 'tapwater_goal', 'basis', 'carcinogenic_effects', 'mutagenic_effects', 'noncancer_effects']
D4a = ['contaminant', 'estuarine_chronic_toxicity', 'estuarine_acute_toxicity', 'freshwater_chronic_toxicity', 'freshwater_acute_toxicity', 'marine_chronic_toxicity', 'marine_acute_toxicity']
D4b = ['contaminant', 'estuarine_goal', 'estuarine_basis', 'freshwater_goal', 'freshwater_basis', 'marine_goal', 'marine_basis']
D4c = ['contaminant', 'estuarine_goal', 'estuarine_basis', 'freshwater_goal', 'freshwater_basis', 'saltwater_goal', 'saltwater_basis']
D4d = ['contaminant', 'freshwater_chronic', 'freshwater_acute', 'saltwater_chronic', 'saltwater_acute']
D4e = ['contaminant', 'freshwater_usepa_chronic', 'freshwater_usepa_acute', 'freshwater_other_chronic', 'freshwater_chronic_basis', 'freshwater_other_acute', 'freshwater_other_basis',  'marine_usepa_chronic', 'marine_usepa_acute', 'marine_other_chronic', 'marine_chronic_basis', 'marine_other_acute', 'marine_other_basis'] 
D4f = ['contaminant', 'criteria', 'basis', 'hi_doh_wqs', 'usepa_nwqc']
D5 = ['contaminant', 'agriculturual_water_goals']
E= ['other', 'contaminant', 'extra', 'koc', 'koc_for_leaching', 'h', 'daf', 'saturation_limit', 'concentration_close_drinking', 'concentration_far_drinking',  'concentration_close_not_drinking', 'concentration_far_not_drinking', 'leaching_close_drinking', 'leaching_far_drinking',  'leaching_close_not_drinking', 'leaching_far_not_drinking',]
F2=['contaminant', 'final_unrestricted_action_level', 'final_commerical_action_level', 'raw_unrestricted_action_level', 'raw_commercial_action_level', 'soil_saturation_limit', 'vp', 'ort_ugm3', 'ort_ppmv',  'odor_index']
G1 = ['contaminant', 'eal', 'basis', 'solubility', 'taste_odor', 'odor_basis', 'upper_limit']
I1 = ['contaminant', 'eal', 'basis', 'carcinogens', 'mutagens', 'noncarcinogens_final', 'noncarcinogens_hq', 'saturation']
I2 = ['contaminant', 'eal', 'basis', 'carcinogens', 'noncarcinogens_final', 'noncarcinogens_hq', 'saturation']
K = ['contaminant', 'range', 'upper_bound', 'background', 'action_level']
L = ['contaminant', 'residential', 'commercial']

columns = {
	9:A,
	10:A,
	11:C,
	12:D,
	13:A1,
	14:A1,
	15:A1,
	16:A1,
	17:C1a,
	18:C1a,
	19:C2,
	20:C3,
	21:D1a,
	22:D1a,
	23:D1c,
	24:D1c,
	25:D2a,
	26:D2b,
	27:D2b,
	28:D3a,
	29:D3b,
	30:D4a,
	31:D4b,
	32:D4c,
	33:D4d,
	34:D4e,
	35:D4f,
	36:D5,
	37:E, 
	38:[], # F-1
	39:F2,
	40:F2,
	41:G1,
	42:G1, # G-2
	43:G1, # G-3
	44:G1, # G-4
	45:[], # H
	46:I1,
	47:I2,
	48:I2,
	49:[], # J
	50:K,
	51:L
	
	
} #sheet:names


###################
# parse arguments #
###################
def _parse_args():
	help_ = """Command line program to read in a single sheet from an excel file and upload to a database. It defaults to creating a sqlite database in the current directory with a provide name appened with ".db", with the default surfer.db
	
	The tables are hard to parse. The column names are not well lined-up, so I had trouble finding a consistent way to programmatically extract them.
	
	This code works on the sheet "Table A-1", sheet 13. I'll test on others later and adjust as needed. The default schema is the schema that works for Table A-1. Eventually, we should program in all the schemas.
	
	The --mongo flag will instead write to your mongodb.
	
	For this to work, mongodb needs to be running on your machine. On my mac, I used homebrew:
	brew install mongodb
I then needed to created the folder:
    /data/db
I could then run mongodb by typing:
    mongod
	"""

	parser = argparse.ArgumentParser(description=help_)
	parser.add_argument('input', help='path to EAL Surfer xlsx file')
	parser.add_argument('--skiprows', type=int, default=6,
						help="Number of nondatarows in the excel file, to include all headers, as the headers don't parse smoothly.")
	parser.add_argument('--names', default=None,
						help='comma-separated list of column names. The names for sheet 13 will default')
	parser.add_argument('--sheet', default=13, type=int,
						help='Sheet number to draw from (0-based). Data tables start with sheet 13.')
	parser.add_argument('--mongohost', default='localhost',
						help='host where mongodb server is running')
	parser.add_argument('--mongoport', default=27017, type=int,
						help='port where mongodb server is running')
	parser.add_argument('--db', default='surfer',
						help='name of  database')
	parser.add_argument('--collection', default='a1',
						help='name of collection/table')
	parser.add_argument('--mongo', action='store_true',
						help='uses mongodb instead of sqllite')
	parser.add_argument('-v', '--verbose', action='store_true')
	args = parser.parse_args()

	if args.names:
		names=arg.names
	else:
		names = [
			"CHEMICAL PARAMETER", "FINAL EAL", "BASIS", "Table F-2",
			"Table L", "Table K", "Table I-1", "Table C-1b", "Table E"
		]
	return args, names
	

####################
# Read excel sheet #
####################

def read_excel(workbook, names, sheet, skiprows, skipfooter):
	df = pandas.read_excel(
		workbook,
        sheetname=sheet,
        header=None,
		skiprows=skiprows,
		skipfooter=skipfooter,
    
    )
	
	df.columns = names 
	
	return df

def get_sheetnames(workbook):
	"""

	"""
	return  {i:name
			 for i,name in enumerate(workbook.sheetnames)
			 if name.startswith('Table') or name.startswith('Summary')}

def get_skiprows(workbook, sheets):
	"""
	Skip the the first row containing ACENAPHTHENE or #ACENAPHTHENE
	Headers can't be automatically derived
	"""
	result = {}
	for sheetnum in sheets:
		if sheetnum == 37:
			col = 1
		else:
			col = 0
		sheet = workbook.worksheets[sheetnum]
		for i,row in enumerate(sheet.rows):
			try:
				if 'ACENAPHTHENE' in row[col].value:
					result[sheetnum]=list(range(i))
					break
			except TypeError:
				pass
	return result

def get_skipfooter(workbook, sheets):
	result = {}
	for sheetnum in sheets:
		if sheetnum == 37:
			col = 1
		else:
			col = 0
		sheet = workbook.worksheets[sheetnum]
		result[sheetnum] = []
		done = False
		for i, row in enumerate(sheet.rows):
			try:
				if 'ZINC' in row[col].value:
					done = True
					continue
				if done:
					result[sheetnum].append(i)
			except TypeError:
				pass
	return {sheet:len(rows) for sheet, rows in result.items()}
	

######################
# Upload to database #
######################

def sqlite_load(df, db_name, table, names, safe=False):

	logging.debug("Loading to sqlite {} {}".format(db_name, table))
	import sqlite3
	conn = sqlite3.connect(db_name)
	c = conn.cursor()

	try:
		create_command='''CREATE TABLE "{}" {}'''.format(table, tuple(names))
		c.execute(create_command)
	except sqlite3.OperationalError:
		# Probably the table exists
		if safe:
			# Stop loading the data
			return
		else:
			# Destroy the table
			c.execute('drop table "{}"'.format(table))
			c.execute(create_command)
		# An alternative is to allow the new data to be inserted
	
	insert_command = 'INSERT INTO "{}" VALUES ({})'.format(table,','.join(['?']*len(names)))
	c.executemany(insert_command, df.itertuples(index=False, name=None))
	conn.commit()

	# To test
	c.execute('SELECT * from "{}"'.format(table))
	logging.debug(c.fetchone())

def mongo_load(df, db_name, collection_name,
			   mongohost='localhost', mongoport=27017):
	"""
	Assumes mongodb is running on machine
	"""
	logging.debug("Loading to mongo {} {}".format(db_name, collection_name))
	from pymongo import MongoClient

	records =  json.loads(df.T.to_json()).values()
	
	client = MongoClient(mongohost, mongoport)
	db = client[db_name]
	collection = db[collection_name]
	collection.insert_many(records) # May create duplicate records

	# To test
	logging.debug(collection.find_one())


def load_sheet_2(workbook, db_name, collection_name, mongo=False, startrow=32, startcol='M', lastcol='Y'):
	"""
	"""
	columns = ['parameter', 'cas', 'contaminant', 'cancer_risk_residential', 'cancer_risk_ci', 'cancer_risk_workers', 'hard_quotient', 'metal', 'volatile', 'persistent', 'modeled_koc', 'code', 'notes']

	df = pandas.read_excel(
		workbook,
        sheetname=2,
        header=startrow-1,
		parse_cols="{}:{}".format(startcol,lastcol),
    )
	df.columns = columns
	logging.debug(df.head())
	if mongo:
		mongo_load(df, db_name, collection_name)
	else:
		sqlite_load(df, db_name, collection_name, columns)


		

	
class Loader:
	"""

	"""
	def __init__(self, workbook_file, db_name, sheets=None, mongo=False, sheet2=True):
		"""
		tables in sheets 9-51
		"""
		self.workbook_file = workbook_file
		if sheets:
			self.sheets = sheets
		else:
			self.sheets = list(range(9,52)) + []
		self.db_name = db_name
		workbook = openpyxl.load_workbook(workbook_file)
		self.collections = get_sheetnames(workbook)
		self.skiprows = get_skiprows(workbook, self.sheets)
		self.skipfooter = get_skipfooter(workbook, self.sheets)
		workbook.close()
		if sheet2:
			load_sheet_2(workbook_file, db_name, "auxillary", mongo)
		self._load(mongo)

		
	def _load(self, mongo):
		"""
		"""
		for sheet in self.sheets:
			logging.info("Loading sheet {}".format(sheet))
			try:
				df = read_excel(self.workbook_file,
								columns[sheet], sheet, self.skiprows[sheet],
								self.skipfooter[sheet]
				)
				logging.debug(df.head())
			
				if mongo: # Use mongodb

					mongo_load(df, self.db_name, self.collections[sheet])

				else: # Use sqlite3
					sqlite_load(df, self.db_name,
								self.collections[sheet],
		   						columns[sheet])			
			except:
				logging.debug("ERROR")
	
	

if __name__ == '__main__':

	args, names = _parse_args()
	if args.verbose:
		logging.root.setLevel(logging.DEBUG)
	Loader(args.input, args.db, mongo=args.mongo)


	# df = read_excel(args.input, names, args.sheet, args.skiprows)
	
	# if args.mongo: # Use mongodb
	# 	mongo_load(df, args.db, args.collection)

	# else: # Use sqlite3
	# 	db_name = args.db+ ".db"
	# 	table_name = "table_" + str(args.sheet)
	# 	sqlite_load(df, db_name, table_name, names)
	
# Finished




